import pypdf
import pdfplumber
import docx
import openpyxl
from typing import List

class DocumentParser:
    def parse_pdf(self, file_path: str) -> str:
        text = ""
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                # Attempt to extract text with layout preservation
                page_text = page.extract_text(layout=True)
                if page_text:
                    text += page_text + "\n"
        return text

    def parse_docx(self, file_path: str) -> str:
        doc = docx.Document(file_path)
        return "\n".join([para.text for para in doc.paragraphs])

    def parse_xlsx(self, file_path: str) -> str:
        wb = openpyxl.load_workbook(file_path) # Simplified
        text = ""
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                text += " ".join([str(cell) for cell in row if cell is not None]) + "\n"
        return text

    def parse(self, file_path: str) -> str:
        if file_path.endswith(".pdf"):
            return self.parse_pdf(file_path)
        elif file_path.endswith(".docx"):
            return self.parse_docx(file_path)
        elif file_path.endswith(".xlsx"):
            return self.parse_xlsx(file_path)
        else:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()
